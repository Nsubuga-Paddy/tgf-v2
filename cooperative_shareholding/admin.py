from django.contrib import admin

from accounts.admin import format_user_autocomplete_label
from core.admin_base import ExportableAdminMixin

from .models import (
    CooperativeGlobalDefaults,
    CooperativeIssuancePeriod,
    CooperativeShareholding,
    DividendAllocationLine,
    DividendChoiceRequest,
    DividendDisbursement,
    ShareAcquisitionLine,
)


class DividendDisbursementInline(admin.TabularInline):
    model = DividendDisbursement
    extra = 0
    fields = (
        "fulfillment_type",
        "amount",
        "shares_count",
        "disbursed_at",
        "payment_reference",
        "notes",
    )
    readonly_fields = (
        "fulfillment_type",
        "amount",
        "shares_count",
        "disbursed_at",
        "notes",
    )
    # payment_reference editable for MoMo/bank refs after manual payout


class DividendAllocationLineInline(admin.TabularInline):
    model = DividendAllocationLine
    extra = 0
    fields = ("action_type", "amount", "shares_count")
    readonly_fields = ("shares_count",)


class ShareAcquisitionLineInline(admin.TabularInline):
    model = ShareAcquisitionLine
    extra = 1
    fields = (
        "receipt_number",
        "acquisition_date",
        "shares_held",
        "share_amount",
        "price_per_share",
        "source_description",
    )


@admin.register(CooperativeIssuancePeriod)
class CooperativeIssuancePeriodAdmin(admin.ModelAdmin):
    list_display = ("name", "usd_to_ugx_rate", "created_at")
    search_fields = ("name",)
    ordering = ("-created_at",)


@admin.register(CooperativeGlobalDefaults)
class CooperativeGlobalDefaultsAdmin(admin.ModelAdmin):
    list_display = (
        "reinvest_share_price",
        "blue_diamond_usd_threshold",
        "updated_at",
    )

    def has_add_permission(self, request):
        return not CooperativeGlobalDefaults.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(CooperativeShareholding)
class CooperativeShareholdingAdmin(ExportableAdminMixin, admin.ModelAdmin):
    list_display = (
        "member_display",
        "total_shares_display",
        "current_share_price",
        "dividend_rate",
        "dividend_election_open",
        "issuance_period",
    )
    list_filter = ("dividend_election_open", "certificate_status")
    search_fields = (
        "user__username",
        "user__first_name",
        "user__last_name",
        "user__profile__account_number",
    )
    inlines = [ShareAcquisitionLineInline]
    autocomplete_fields = ("user", "issuance_period")

    def get_autocomplete_label(self, obj):
        return format_user_autocomplete_label(obj.user)

    fieldsets = (
        (
            "Member",
            {"fields": ("user", "year_joined", "certificate_status", "admin_notes")},
        ),
        (
            "Dividend issuance (set per member)",
            {
                "fields": (
                    "current_share_price",
                    "dividend_rate",
                    "dividend_election_open",
                    "issuance_period",
                ),
                "description": (
                    "Current share price and dividend rate apply to this member. "
                    "Issuance period supplies the USD→UGX rate for Blue Diamond tier."
                ),
            },
        ),
    )

    def member_display(self, obj):
        return format_user_autocomplete_label(obj.user)

    member_display.short_description = "Member"
    member_display.admin_order_field = "user__last_name"

    def total_shares_display(self, obj):
        return obj.total_shares

    total_shares_display.short_description = "Total shares"


@admin.register(DividendChoiceRequest)
class DividendChoiceRequestAdmin(ExportableAdminMixin, admin.ModelAdmin):
    list_display = (
        "user_full_name",
        "username_display",
        "phone_display",
        "email_display",
        "total_dividend_display",
        "allocation_summary_display",
        "member_notes_preview",
        "bank_details_display",
        "status",
        "created_at",
    )
    list_filter = ("status", "created_at")
    autocomplete_fields = ("shareholding",)
    search_fields = (
        "shareholding__user__username",
        "shareholding__user__first_name",
        "shareholding__user__last_name",
        "shareholding__user__email",
        "shareholding__user__profile__whatsapp_number",
        "shareholding__user__profile__account_number",
        "shareholding__user__profile__bank_name",
        "shareholding__user__profile__bank_account_number",
        "shareholding__user__profile__bank_account_name",
        "member_notes",
    )
    readonly_fields = (
        "user_full_name",
        "username_display",
        "phone_display",
        "email_display",
        "bank_details_display",
        "allocation_summary_display",
        "ledger_applied_at",
        "created_at",
        "processed_at",
    )

    def get_readonly_fields(self, request, obj=None):
        ro = list(super().get_readonly_fields(request, obj))
        if obj:
            ro.extend(["shareholding", "total_dividend"])
        return ro
    inlines = (DividendAllocationLineInline, DividendDisbursementInline)
    date_hierarchy = "created_at"
    fieldsets = (
        (
            "Member (for payments & contact)",
            {
                "fields": (
                    "user_full_name",
                    "username_display",
                    "email_display",
                    "phone_display",
                    "bank_details_display",
                ),
            },
        ),
        (
            "Dividend request",
            {
                "fields": (
                    "shareholding",
                    "total_dividend",
                    "allocation_summary_display",
                    "member_notes",
                    "status",
                ),
            },
        ),
        (
            "Admin",
            {
                "fields": ("admin_notes", "processed_at", "ledger_applied_at"),
            },
        ),
        (
            "Timestamps",
            {
                "fields": ("created_at",),
                "classes": ("collapse",),
            },
        ),
    )

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related(
                "shareholding",
                "shareholding__user",
                "shareholding__user__profile",
            )
            .prefetch_related("allocation_lines")
        )

    def _user(self, obj):
        return obj.shareholding.user

    def _profile(self, obj):
        return getattr(obj.shareholding.user, "profile", None)

    def user_full_name(self, obj):
        user = self._user(obj)
        full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
        return full_name or user.get_username()

    user_full_name.short_description = "Full name"
    user_full_name.admin_order_field = "shareholding__user__last_name"

    def username_display(self, obj):
        return self._user(obj).get_username()

    username_display.short_description = "Username"
    username_display.admin_order_field = "shareholding__user__username"

    def email_display(self, obj):
        return self._user(obj).email or "—"

    email_display.short_description = "Email"
    email_display.admin_order_field = "shareholding__user__email"

    def phone_display(self, obj):
        profile = self._profile(obj)
        if profile and profile.whatsapp_number:
            return str(profile.whatsapp_number)
        return "—"

    phone_display.short_description = "Phone"

    def bank_details_display(self, obj):
        profile = self._profile(obj)
        if not profile:
            return "—"
        parts = []
        if profile.bank_name:
            parts.append(profile.bank_name)
        if profile.bank_account_number:
            parts.append(profile.bank_account_number)
        if profile.bank_account_name:
            parts.append(profile.bank_account_name)
        return " | ".join(parts) if parts else "Not provided"

    bank_details_display.short_description = "Bank account"

    def total_dividend_display(self, obj):
        return f"UGX {obj.total_dividend:,.0f}"

    total_dividend_display.short_description = "Total dividend"
    total_dividend_display.admin_order_field = "total_dividend"

    def member_notes_preview(self, obj):
        if not obj.member_notes:
            return "—"
        text = obj.member_notes.strip()
        return text[:60] + "…" if len(text) > 60 else text

    member_notes_preview.short_description = "Member notes"

    def allocation_summary_display(self, obj):
        return obj.allocation_summary

    allocation_summary_display.short_description = "Allocation"

    def _allocation_amounts(self, obj):
        from decimal import Decimal

        from .models import DividendAllocationLine

        amounts = {choice[0]: Decimal("0") for choice in DividendAllocationLine.ActionType.choices}
        for line in obj.allocation_lines.all():
            amounts[line.action_type] = line.amount
        return amounts

    def _get_dividend_request_export_rows(self, queryset):
        from .models import DividendAllocationLine

        headers = [
            "Full name",
            "Username",
            "Email",
            "Phone",
            "Account number",
            "Total dividend (UGX)",
            "Cash (UGX)",
            "MCS shares (UGX)",
            "MESU shares (UGX)",
            "Savings (UGX)",
            "Allocation summary",
            "Member notes",
            "Bank name",
            "Bank account number",
            "Bank account name",
            "Status",
            "Submitted at",
            "Admin notes",
        ]
        rows = []
        for sub in queryset:
            user = sub.shareholding.user
            profile = getattr(user, "profile", None)
            amounts = self._allocation_amounts(sub)
            rows.append(
                [
                    self.user_full_name(sub),
                    user.get_username(),
                    user.email or "",
                    str(profile.whatsapp_number) if profile and profile.whatsapp_number else "",
                    profile.account_number if profile else "",
                    f"{sub.total_dividend:,.0f}",
                    f"{amounts[DividendAllocationLine.ActionType.CASH]:,.0f}",
                    f"{amounts[DividendAllocationLine.ActionType.MCS_SHARES]:,.0f}",
                    f"{amounts[DividendAllocationLine.ActionType.MESU_SHARES]:,.0f}",
                    f"{amounts[DividendAllocationLine.ActionType.SAVINGS]:,.0f}",
                    sub.allocation_summary,
                    sub.member_notes or "",
                    profile.bank_name if profile else "",
                    profile.bank_account_number if profile else "",
                    profile.bank_account_name if profile else "",
                    sub.get_status_display(),
                    sub.created_at.strftime("%Y-%m-%d %H:%M:%S") if sub.created_at else "",
                    sub.admin_notes or "",
                ]
            )
        return headers, rows

    def get_actions(self, request):
        actions = super().get_actions(request)

        def export_csv_wrapper(modeladmin, req, queryset):
            return self.export_dividend_requests_csv(req, queryset)

        export_csv_wrapper.short_description = (
            "Export dividend requests for payments (CSV)"
        )

        def export_excel_wrapper(modeladmin, req, queryset):
            return self.export_dividend_requests_excel(req, queryset)

        export_excel_wrapper.short_description = (
            "Export dividend requests for payments (Excel)"
        )

        actions["export_dividend_requests_csv"] = (
            export_csv_wrapper,
            "export_dividend_requests_csv",
            export_csv_wrapper.short_description,
        )
        actions["export_dividend_requests_excel"] = (
            export_excel_wrapper,
            "export_dividend_requests_excel",
            export_excel_wrapper.short_description,
        )
        return actions

    def export_dividend_requests_csv(self, request, queryset):
        import csv
        from datetime import datetime

        from django.http import HttpResponse

        qs = self.get_queryset(request).filter(pk__in=queryset.values_list("pk", flat=True))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="dividend_request_submissions_{timestamp}.csv"'
        )
        response.write("\ufeff")
        writer = csv.writer(response)
        headers, rows = self._get_dividend_request_export_rows(qs)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    def export_dividend_requests_excel(self, request, queryset):
        from datetime import datetime
        from io import BytesIO

        from django.http import HttpResponse

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return self.export_dividend_requests_csv(request, queryset)

        qs = self.get_queryset(request).filter(pk__in=queryset.values_list("pk", flat=True))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        wb = Workbook()
        ws = wb.active
        ws.title = "Dividend requests"
        headers, rows = self._get_dividend_request_export_rows(qs)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="dividend_request_submissions_{timestamp}.xlsx"'
        )
        return response

    def save_model(self, request, obj, form, change):
        previous_status = None
        if change and obj.pk:
            previous_status = (
                DividendChoiceRequest.objects.filter(pk=obj.pk)
                .values_list("status", flat=True)
                .first()
            )
        super().save_model(request, obj, form, change)
        newly_approved = obj.status in (
            DividendChoiceRequest.Status.APPROVED,
            DividendChoiceRequest.Status.PROCESSED,
        ) and previous_status not in (
            DividendChoiceRequest.Status.APPROVED,
            DividendChoiceRequest.Status.PROCESSED,
        )
        if newly_approved:
            from .services import apply_approved_dividend_ledger

            submission = DividendChoiceRequest.objects.prefetch_related(
                "allocation_lines", "shareholding"
            ).get(pk=obj.pk)
            apply_approved_dividend_ledger(submission)
            from django.contrib import messages as admin_messages

            self.message_user(
                request,
                "Dividend disbursements recorded for the member. MCS reinvestments added to share acquisitions; cash/savings show on their dividend statement.",
                level=admin_messages.SUCCESS,
            )
