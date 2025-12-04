"""
Admin Export Utilities
Provides CSV, Excel, and PDF export functionality for Django admin.
"""
import csv
from datetime import datetime
from io import BytesIO
from typing import List, Any, Optional

from django.http import HttpResponse
from django.utils.text import slugify


def export_to_csv(modeladmin, request, queryset, filename: str = None, fields: List[str] = None):
    """
    Export queryset to CSV file.
    
    Args:
        modeladmin: The ModelAdmin instance
        request: The HTTP request
        queryset: The queryset to export
        filename: Optional filename (auto-generated if not provided)
        fields: Optional list of field names to export (uses list_display if not provided)
    """
    opts = modeladmin.model._meta
    
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{slugify(opts.verbose_name_plural)}_{timestamp}.csv"
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')  # UTF-8 BOM for Excel compatibility
    
    writer = csv.writer(response)
    
    # Determine which fields to export
    if fields:
        field_names = fields
    else:
        # Use list_display from admin, excluding action_checkbox
        field_names = [f for f in modeladmin.list_display if f != 'action_checkbox']
    
    # Write header row
    headers = []
    for field_name in field_names:
        if hasattr(modeladmin, field_name):
            # Custom admin method
            method = getattr(modeladmin, field_name)
            if hasattr(method, 'short_description'):
                headers.append(method.short_description)
            else:
                headers.append(field_name.replace('_', ' ').title())
        else:
            # Model field
            try:
                field = opts.get_field(field_name)
                headers.append(field.verbose_name.title())
            except:
                headers.append(field_name.replace('_', ' ').title())
    
    writer.writerow(headers)
    
    # Write data rows
    for obj in queryset:
        row = []
        for field_name in field_names:
            if hasattr(modeladmin, field_name):
                # Custom admin method
                method = getattr(modeladmin, field_name)
                value = method(obj)
                # Clean HTML tags from value
                if isinstance(value, str):
                    import re
                    value = re.sub('<[^<]+?>', '', value)
            else:
                # Model field
                try:
                    value = getattr(obj, field_name)
                    # Handle foreign keys
                    if hasattr(value, '__str__'):
                        value = str(value)
                except AttributeError:
                    value = ''
            
            row.append(value)
        writer.writerow(row)
    
    return response


def export_to_excel(modeladmin, request, queryset, filename: str = None, fields: List[str] = None):
    """
    Export queryset to Excel file (.xlsx).
    
    Args:
        modeladmin: The ModelAdmin instance
        request: The HTTP request
        queryset: The queryset to export
        filename: Optional filename (auto-generated if not provided)
        fields: Optional list of field names to export (uses list_display if not provided)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return export_to_csv(modeladmin, request, queryset, filename, fields)
    
    opts = modeladmin.model._meta
    
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{slugify(opts.verbose_name_plural)}_{timestamp}.xlsx"
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = opts.verbose_name_plural[:31]  # Excel sheet name limit
    
    # Determine which fields to export
    if fields:
        field_names = fields
    else:
        field_names = [f for f in modeladmin.list_display if f != 'action_checkbox']
    
    # Write header row with styling
    headers = []
    for field_name in field_names:
        if hasattr(modeladmin, field_name):
            method = getattr(modeladmin, field_name)
            if hasattr(method, 'short_description'):
                headers.append(method.short_description)
            else:
                headers.append(field_name.replace('_', ' ').title())
        else:
            try:
                field = opts.get_field(field_name)
                headers.append(field.verbose_name.title())
            except:
                headers.append(field_name.replace('_', ' ').title())
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(field_names, 1):
            if hasattr(modeladmin, field_name):
                method = getattr(modeladmin, field_name)
                value = method(obj)
                # Clean HTML tags
                if isinstance(value, str):
                    import re
                    value = re.sub('<[^<]+?>', '', value)
            else:
                try:
                    value = getattr(obj, field_name)
                    if hasattr(value, '__str__'):
                        value = str(value)
                except AttributeError:
                    value = ''
            
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = len(headers[col_num - 1])
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f'{column_letter}{row}'].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_to_pdf(modeladmin, request, queryset, filename: str = None, fields: List[str] = None, 
                  title: str = None, orientation: str = 'landscape'):
    """
    Export queryset to PDF file.
    
    Args:
        modeladmin: The ModelAdmin instance
        request: The HTTP request
        queryset: The queryset to export
        filename: Optional filename (auto-generated if not provided)
        fields: Optional list of field names to export (uses list_display if not provided)
        title: Optional title for the PDF
        orientation: 'portrait' or 'landscape' (default: 'landscape')
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4, landscape as rl_landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        # Fallback to CSV if reportlab not installed
        return export_to_csv(modeladmin, request, queryset, filename, fields)
    
    opts = modeladmin.model._meta
    
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{slugify(opts.verbose_name_plural)}_{timestamp}.pdf"
    
    if not title:
        title = opts.verbose_name_plural.title()
    
    # Create PDF
    buffer = BytesIO()
    pagesize = rl_landscape(A4) if orientation == 'landscape' else A4
    doc = SimpleDocTemplate(buffer, pagesize=pagesize)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.25 * inch))
    
    # Determine which fields to export
    if fields:
        field_names = fields
    else:
        field_names = [f for f in modeladmin.list_display if f != 'action_checkbox']
    
    # Prepare table data
    headers = []
    for field_name in field_names:
        if hasattr(modeladmin, field_name):
            method = getattr(modeladmin, field_name)
            if hasattr(method, 'short_description'):
                headers.append(method.short_description)
            else:
                headers.append(field_name.replace('_', ' ').title())
        else:
            try:
                field = opts.get_field(field_name)
                headers.append(field.verbose_name.title())
            except:
                headers.append(field_name.replace('_', ' ').title())
    
    data = [headers]
    
    # Add data rows
    for obj in queryset[:100]:  # Limit to 100 rows for PDF
        row = []
        for field_name in field_names:
            if hasattr(modeladmin, field_name):
                method = getattr(modeladmin, field_name)
                value = method(obj)
                # Clean HTML tags
                if isinstance(value, str):
                    import re
                    value = re.sub('<[^<]+?>', '', value)
                    # Truncate long values
                    if len(value) > 50:
                        value = value[:47] + '...'
            else:
                try:
                    value = getattr(obj, field_name)
                    if hasattr(value, '__str__'):
                        value = str(value)
                    # Truncate long values
                    if isinstance(value, str) and len(value) > 50:
                        value = value[:47] + '...'
                except AttributeError:
                    value = ''
            
            row.append(value)
        data.append(row)
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Add footer with metadata
    elements.append(Spacer(1, 0.25 * inch))
    footer_text = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {queryset.count()}"
    elements.append(Paragraph(footer_text, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def create_export_actions(model_name: str):
    """
    Factory function to create CSV, Excel, and PDF export actions for a model.
    
    Usage in admin.py:
        from core.admin_exports import create_export_actions
        
        @admin.register(MyModel)
        class MyModelAdmin(admin.ModelAdmin):
            actions = create_export_actions('MyModel')
    
    Returns:
        List of action functions
    """
    
    def export_as_csv(modeladmin, request, queryset):
        """Export selected items as CSV"""
        return export_to_csv(modeladmin, request, queryset)
    export_as_csv.short_description = f"Export selected {model_name} as CSV"
    
    def export_as_excel(modeladmin, request, queryset):
        """Export selected items as Excel"""
        return export_to_excel(modeladmin, request, queryset)
    export_as_excel.short_description = f"Export selected {model_name} as Excel"
    
    def export_as_pdf(modeladmin, request, queryset):
        """Export selected items as PDF"""
        return export_to_pdf(modeladmin, request, queryset)
    export_as_pdf.short_description = f"Export selected {model_name} as PDF"
    
    return [export_as_csv, export_as_excel, export_as_pdf]

