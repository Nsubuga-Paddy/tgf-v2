from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.models import User
from .models import UserProfile, Project, AccountNumberCounter, WithdrawalRequest, GWCContribution, MESUInterest
from core.admin_base import ExportableAdminMixin


class UserProfileInline(admin.StackedInline):
    model = UserProfile
    can_delete = False
    verbose_name_plural = 'Profile'
    fields = (
        'photo', 'whatsapp_number', 'national_id', 'birthdate', 
        'address', 'bio', 'account_number', 'bank_name', 'bank_account_number', 
        'bank_account_name', 'is_verified', 'is_admin', 'projects'
    )
    readonly_fields = ('account_number', 'created_at', 'updated_at')
    filter_horizontal = ('projects',)
    
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        # Ensure the projects field uses the proper ManyToMany widget
        if 'projects' in form.base_fields:
            form.base_fields['projects'].widget.can_add_related = True
            form.base_fields['projects'].widget.can_change_related = True
            form.base_fields['projects'].widget.can_delete_related = False
        return form


class UserAdmin(BaseUserAdmin):
    inlines = (UserProfileInline,)
    list_display = ('username', 'email', 'first_name', 'last_name', 'is_staff', 'get_account_number', 'get_verification_status')
    list_filter = ('is_staff', 'is_superuser', 'is_active', 'profile__is_verified', 'profile__is_admin')
    search_fields = ('username', 'first_name', 'last_name', 'email', 'profile__account_number')
    ordering = ('username',)
    
    def get_account_number(self, obj):
        if hasattr(obj, 'profile'):
            return obj.profile.account_number
        return 'No Profile'
    get_account_number.short_description = 'Account Number'
    
    def get_verification_status(self, obj):
        if hasattr(obj, 'profile'):
            if obj.profile.is_verified:
                return '✅ Verified'
            return '⏳ Pending'
        return '❌ No Profile'
    get_verification_status.short_description = 'Status'


@admin.register(UserProfile)
class UserProfileAdmin(ExportableAdminMixin, admin.ModelAdmin):
    list_display = ('user', 'account_number', 'whatsapp_number', 'get_bank_info', 'is_verified', 'is_admin', 'get_projects', 'created_at')
    list_filter = ('is_verified', 'is_admin', 'projects', 'created_at', 'bank_name')
    search_fields = ('user__username', 'user__first_name', 'user__last_name', 'account_number', 'whatsapp_number', 'bank_name', 'bank_account_number', 'bank_account_name')
    readonly_fields = ('account_number', 'created_at', 'updated_at')
    filter_horizontal = ('projects',)
    fieldsets = (
        ('User Information', {
            'fields': ('user', 'photo', 'account_number')
        }),
        ('Personal Details', {
            'fields': ('whatsapp_number', 'national_id', 'birthdate', 'address', 'bio')
        }),
        ('Bank Account Information', {
            'fields': ('bank_name', 'bank_account_number', 'bank_account_name'),
            'description': 'Bank account details used for withdrawals and payments. Admin will use these to send money to users.'
        }),
        ('Status & Permissions', {
            'fields': ('is_verified', 'is_admin', 'projects')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user').prefetch_related('projects')
    
    def get_projects(self, obj):
        """Display projects as a comma-separated list"""
        projects = obj.projects.all()
        if projects:
            return ', '.join([project.name for project in projects])
        return 'No projects'
    get_projects.short_description = 'Projects'
    
    def get_bank_info(self, obj):
        """Display bank account information"""
        if obj.bank_name and obj.bank_account_number:
            return f"{obj.bank_name} - {obj.bank_account_number}"
        return 'Not provided'
    get_bank_info.short_description = 'Bank Account'
    
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        # Ensure the projects field uses the proper ManyToMany widget
        if 'projects' in form.base_fields:
            form.base_fields['projects'].widget.can_add_related = True
            form.base_fields['projects'].widget.can_change_related = True
            form.base_fields['projects'].widget.can_delete_related = False
        return form
    
    def get_actions(self, request):
        """Add custom export actions for 52WSC users"""
        actions = super().get_actions(request)
        
        # Wrap instance methods to match Django's expected signature (modeladmin, request, queryset)
        def export_52wsc_csv_wrapper(modeladmin, request, queryset):
            return self.export_52wsc_users_csv(request, queryset)
        export_52wsc_csv_wrapper.short_description = "Export 52WSC users report (CSV)"
        
        def export_52wsc_excel_wrapper(modeladmin, request, queryset):
            return self.export_52wsc_users_excel(request, queryset)
        export_52wsc_excel_wrapper.short_description = "Export 52WSC users report (Excel)"
        
        def export_52wsc_pdf_wrapper(modeladmin, request, queryset):
            return self.export_52wsc_users_pdf(request, queryset)
        export_52wsc_pdf_wrapper.short_description = "Export 52WSC users report (PDF)"
        
        actions['export_52wsc_users_csv'] = (
            export_52wsc_csv_wrapper,
            'export_52wsc_users_csv',
            export_52wsc_csv_wrapper.short_description
        )
        actions['export_52wsc_users_excel'] = (
            export_52wsc_excel_wrapper,
            'export_52wsc_users_excel',
            export_52wsc_excel_wrapper.short_description
        )
        actions['export_52wsc_users_pdf'] = (
            export_52wsc_pdf_wrapper,
            'export_52wsc_users_pdf',
            export_52wsc_pdf_wrapper.short_description
        )
        
        return actions
    
    def _get_52wsc_users_data(self):
        """Helper method to get 52WSC users data"""
        # Get all users with the 52 Weeks Saving Challenge project
        try:
            project = Project.objects.get(name='52 Weeks Saving Challenge')
            profiles = UserProfile.objects.filter(projects=project).select_related('user').order_by('user__last_name', 'user__first_name')
        except Project.DoesNotExist:
            profiles = UserProfile.objects.none()
        
        headers = [
            'Full Name',
            'Amount Saved',
            'Interest Earned',
            'Total Savings and Interest',
            'Account Number',
            'Phone Number'
        ]
        
        data = []
        for profile in profiles:
            user = profile.user
            full_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.get_username()
            amount_saved = profile.get_amount_saved()
            interest_earned = profile.get_total_interest_earned()
            total_savings = profile.get_total_savings()
            
            row = [
                full_name,
                f"{amount_saved:,.2f}",
                f"{interest_earned:,.2f}",
                f"{total_savings:,.2f}",
                profile.account_number or '',
                str(profile.whatsapp_number) if profile.whatsapp_number else ''
            ]
            data.append(row)
        
        return headers, data
    
    def export_52wsc_users_csv(self, request, queryset):
        """
        Export all 52WSC users as CSV with their savings and interest data.
        Note: queryset parameter is ignored - this exports ALL 52WSC users.
        """
        import csv
        from django.http import HttpResponse
        from datetime import datetime
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"52wsc_users_report_{timestamp}.csv"
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('\ufeff')  # UTF-8 BOM for Excel compatibility
        
        writer = csv.writer(response)
        headers, data = self._get_52wsc_users_data()
        
        writer.writerow(headers)
        for row in data:
            writer.writerow(row)
        
        return response
    
    export_52wsc_users_csv.short_description = "Export 52WSC users report (CSV)"
    
    def export_52wsc_users_excel(self, request, queryset):
        """
        Export all 52WSC users as Excel with their savings and interest data.
        Note: queryset parameter is ignored - this exports ALL 52WSC users.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            return self.export_52wsc_users_csv(request, queryset)
        
        from django.http import HttpResponse
        from datetime import datetime
        from io import BytesIO
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"52wsc_users_report_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "52WSC Users Report"
        
        headers, data = self._get_52wsc_users_data()
        
        # Write header row with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data rows
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(headers[col_num - 1])
            for row in range(2, ws.max_row + 1):
                cell_value = ws[f'{column_letter}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    export_52wsc_users_excel.short_description = "Export 52WSC users report (Excel)"
    
    def export_52wsc_users_pdf(self, request, queryset):
        """
        Export all 52WSC users as PDF with their savings and interest data.
        Note: queryset parameter is ignored - this exports ALL 52WSC users.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape as rl_landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            # Fallback to CSV if reportlab not installed
            return self.export_52wsc_users_csv(request, queryset)
        
        from django.http import HttpResponse
        from datetime import datetime
        from io import BytesIO
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"52wsc_users_report_{timestamp}.pdf"
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=rl_landscape(A4))
        elements = []
        
        # Add title
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        elements.append(Paragraph("52 Weeks Saving Challenge - Users Report", title_style))
        elements.append(Spacer(1, 0.25 * inch))
        
        headers, data = self._get_52wsc_users_data()
        
        # Prepare table data
        table_data = [headers]
        # Limit to 100 rows for PDF performance
        for row in data[:100]:
            # Truncate long values for PDF display
            truncated_row = []
            for value in row:
                if isinstance(value, str) and len(value) > 40:
                    truncated_row.append(value[:37] + '...')
                else:
                    truncated_row.append(value)
            table_data.append(truncated_row)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Add footer with metadata
        elements.append(Spacer(1, 0.25 * inch))
        total_count = len(data)
        displayed_count = min(total_count, 100)
        footer_text = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {total_count}"
        if displayed_count < total_count:
            footer_text += f" | Displayed: {displayed_count} (PDF limited to 100 rows)"
        elements.append(Paragraph(footer_text, styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    export_52wsc_users_pdf.short_description = "Export 52WSC users report (PDF)"


@admin.register(Project)
class ProjectAdmin(ExportableAdminMixin, admin.ModelAdmin):
    list_display = ('name', 'description', 'get_member_count')
    search_fields = ('name', 'description')
    
    def get_member_count(self, obj):
        return obj.members.count()
    get_member_count.short_description = 'Members'
    
    fieldsets = (
        ('Project Information', {
            'fields': ('name', 'description')
        }),
    )


@admin.register(AccountNumberCounter)
class AccountNumberCounterAdmin(admin.ModelAdmin):
    list_display = ('id', 'created_at')
    readonly_fields = ('created_at',)
    
    def has_add_permission(self, request):
        # Only allow one counter instance
        return not AccountNumberCounter.objects.exists()
    
    def has_delete_permission(self, request, obj=None):
        # Prevent deletion of the counter
        return False


# Re-register UserAdmin
admin.site.unregister(User)
admin.site.register(User, UserAdmin)


@admin.register(WithdrawalRequest)
class WithdrawalRequestAdmin(ExportableAdminMixin, admin.ModelAdmin):
    list_display = ('user_profile', 'amount', 'status', 'created_at', 'get_bank_info')
    list_filter = ('status', 'created_at')
    search_fields = ('user_profile__user__username', 'user_profile__user__first_name', 'user_profile__user__last_name', 'user_profile__account_number')
    readonly_fields = ('created_at', 'updated_at', 'get_bank_info')
    fieldsets = (
        ('Request Information', {
            'fields': ('user_profile', 'amount', 'reason', 'status')
        }),
        ('Bank Account Details', {
            'fields': ('get_bank_info',),
            'description': 'Bank account information from user profile'
        }),
        ('Admin Actions', {
            'fields': ('admin_notes', 'processed_at')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def get_bank_info(self, obj):
        """Display bank account information from user profile"""
        profile = obj.user_profile
        if profile.bank_name and profile.bank_account_number:
            account_name = profile.bank_account_name or 'N/A'
            return f"{profile.bank_name} | {profile.bank_account_number} | {account_name}"
        return 'Not provided'
    get_bank_info.short_description = 'Bank Account'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user_profile__user')
    
    def get_actions(self, request):
        """Add custom export actions along with mixin actions"""
        actions = super().get_actions(request)
        
        # Add custom export actions - wrap instance methods to match Django's expected signature (modeladmin, request, queryset)
        def export_csv_wrapper(modeladmin, request, queryset):
            return self.export_withdrawal_requests_csv(request, queryset)
        export_csv_wrapper.short_description = "Export withdrawal requests with full details (CSV)"
        
        def export_excel_wrapper(modeladmin, request, queryset):
            return self.export_withdrawal_requests_excel(request, queryset)
        export_excel_wrapper.short_description = "Export withdrawal requests with full details (Excel)"
        
        def export_pdf_wrapper(modeladmin, request, queryset):
            return self.export_withdrawal_requests_pdf(request, queryset)
        export_pdf_wrapper.short_description = "Export withdrawal requests with full details (PDF)"
        
        actions['export_withdrawal_requests_csv'] = (
            export_csv_wrapper,
            'export_withdrawal_requests_csv',
            export_csv_wrapper.short_description
        )
        actions['export_withdrawal_requests_excel'] = (
            export_excel_wrapper,
            'export_withdrawal_requests_excel',
            export_excel_wrapper.short_description
        )
        actions['export_withdrawal_requests_pdf'] = (
            export_pdf_wrapper,
            'export_withdrawal_requests_pdf',
            export_pdf_wrapper.short_description
        )
        
        return actions
    
    def _get_withdrawal_data(self, queryset):
        """Helper method to get withdrawal data in a consistent format"""
        headers = [
            'Username',
            'First Name',
            'Last Name',
            'Phone Number',
            'Amount Saved',
            'Interest Earned',
            'Total Savings and Interest',
            'Amount Requested to Withdraw',
            'Bank Name',
            'Bank Account Number',
            'Bank Account Name',
            'Status',
            'Request Date',
            'Reason'
        ]
        
        data = []
        for withdrawal in queryset:
            profile = withdrawal.user_profile
            user = profile.user
            
            # Get financial information
            amount_saved = profile.get_amount_saved()
            interest_earned = profile.get_total_interest_earned()
            total_savings = profile.get_total_savings()
            
            row = [
                user.username or '',
                user.first_name or '',
                user.last_name or '',
                str(profile.whatsapp_number) if profile.whatsapp_number else '',
                f"{amount_saved:,.2f}",
                f"{interest_earned:,.2f}",
                f"{total_savings:,.2f}",
                f"{withdrawal.amount:,.2f}",
                profile.bank_name or '',
                profile.bank_account_number or '',
                profile.bank_account_name or '',
                withdrawal.get_status_display(),
                withdrawal.created_at.strftime('%Y-%m-%d %H:%M:%S') if withdrawal.created_at else '',
                withdrawal.reason or ''
            ]
            data.append(row)
        
        return headers, data
    
    def export_withdrawal_requests_csv(self, request, queryset):
        """
        Export withdrawal requests as CSV with all user and financial details.
        """
        import csv
        from django.http import HttpResponse
        from datetime import datetime
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"withdrawal_requests_{timestamp}.csv"
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('\ufeff')  # UTF-8 BOM for Excel compatibility
        
        writer = csv.writer(response)
        headers, data = self._get_withdrawal_data(queryset)
        
        writer.writerow(headers)
        for row in data:
            writer.writerow(row)
        
        return response
    
    export_withdrawal_requests_csv.short_description = "Export withdrawal requests with full details (CSV)"
    
    def export_withdrawal_requests_excel(self, request, queryset):
        """
        Export withdrawal requests as Excel with all user and financial details.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            return self.export_withdrawal_requests_csv(request, queryset)
        
        from django.http import HttpResponse
        from datetime import datetime
        from io import BytesIO
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"withdrawal_requests_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Withdrawal Requests"
        
        headers, data = self._get_withdrawal_data(queryset)
        
        # Write header row with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data rows
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(headers[col_num - 1])
            for row in range(2, ws.max_row + 1):
                cell_value = ws[f'{column_letter}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    export_withdrawal_requests_excel.short_description = "Export withdrawal requests with full details (Excel)"
    
    def export_withdrawal_requests_pdf(self, request, queryset):
        """
        Export withdrawal requests as PDF with all user and financial details.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape as rl_landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            # Fallback to CSV if reportlab not installed
            return self.export_withdrawal_requests_csv(request, queryset)
        
        from django.http import HttpResponse
        from datetime import datetime
        from io import BytesIO
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"withdrawal_requests_{timestamp}.pdf"
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=rl_landscape(A4))
        elements = []
        
        # Add title
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        elements.append(Paragraph("Withdrawal Requests Report", title_style))
        elements.append(Spacer(1, 0.25 * inch))
        
        headers, data = self._get_withdrawal_data(queryset)
        
        # Prepare table data
        table_data = [headers]
        # Limit to 100 rows for PDF performance
        for row in data[:100]:
            # Truncate long values for PDF display
            truncated_row = []
            for value in row:
                if isinstance(value, str) and len(value) > 40:
                    truncated_row.append(value[:37] + '...')
                else:
                    truncated_row.append(value)
            table_data.append(truncated_row)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Add footer with metadata
        elements.append(Spacer(1, 0.25 * inch))
        total_count = queryset.count()
        displayed_count = min(total_count, 100)
        footer_text = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {total_count}"
        if displayed_count < total_count:
            footer_text += f" | Displayed: {displayed_count} (PDF limited to 100 rows)"
        elements.append(Paragraph(footer_text, styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    export_withdrawal_requests_pdf.short_description = "Export withdrawal requests with full details (PDF)"


@admin.register(GWCContribution)
class GWCContributionAdmin(ExportableAdminMixin, admin.ModelAdmin):
    list_display = ('user_profile', 'amount', 'group_type', 'status', 'created_at')
    list_filter = ('status', 'group_type', 'created_at')
    search_fields = ('user_profile__user__username', 'user_profile__user__first_name', 'user_profile__user__last_name', 'user_profile__account_number')
    readonly_fields = ('created_at', 'updated_at')
    fieldsets = (
        ('Contribution Information', {
            'fields': ('user_profile', 'amount', 'group_type', 'status')
        }),
        ('Admin Actions', {
            'fields': ('admin_notes', 'processed_at')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user_profile__user')


@admin.register(MESUInterest)
class MESUInterestAdmin(ExportableAdminMixin, admin.ModelAdmin):
    list_display = ('user_profile', 'investment_amount', 'number_of_shares', 'status', 'created_at')
    list_filter = ('status', 'created_at')
    search_fields = ('user_profile__user__username', 'user_profile__user__first_name', 'user_profile__user__last_name', 'user_profile__account_number')
    readonly_fields = ('number_of_shares', 'created_at', 'updated_at')
    fieldsets = (
        ('Investment Information', {
            'fields': ('user_profile', 'investment_amount', 'number_of_shares', 'notes', 'status')
        }),
        ('Admin Actions', {
            'fields': ('admin_notes', 'processed_at')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user_profile__user')
